import wikipedia
from openpyxl import Workbook
import openpyxl
import getpass
from openpyxl.chart import BarChart, Reference



def startWikiWordCount():
    #create dictionary with placeholders
    dic = {"Year": "Word Count", 2010: "placeholder", 2010: "placeholder", 2011: "placeholder", 2012: "placeholder"
    , 2013: "placeholder", 2014: "placeholder", 2015: "placeholder", 2016: "placeholder",
    2017: "placeholder", 2018: "placeholder", 2019: "placeholder", 2020: "placeholder",
    2021: "placeholder", 2022: "placeholder"}

    #create wiki pages for each year
    page2010 = wikipedia.page(pageid=43226)
    page2011 = wikipedia.page(pageid=36225)
    page2012 = wikipedia.page(pageid=47374)
    page2013 = wikipedia.page(pageid=46945)
    page2014 = wikipedia.page(pageid=48630)
    page2015 = wikipedia.page(pageid=49708)
    page2016 = wikipedia.page(pageid=51387)
    page2017 = wikipedia.page(pageid=51389)
    page2018 = wikipedia.page(pageid=51390)
    page2019 = wikipedia.page(pageid=51391)
    page2020 = wikipedia.page(pageid=51396)
    page2021 = wikipedia.page(pageid=51397)
    page2022 = wikipedia.page(pageid=52412)

    #create list that will have word counts
    lst = []

    #create word counts for each year
    WC2010 = len(page2010.content.split())
    WC2011 = len(page2011.content.split())
    WC2012 = len(page2012.content.split())
    WC2013 = len(page2013.content.split())
    WC2014 = len(page2014.content.split())
    WC2015 = len(page2015.content.split())
    WC2016 = len(page2016.content.split())
    WC2017 = len(page2017.content.split())
    WC2018 = len(page2018.content.split())
    WC2019 = len(page2019.content.split())
    WC2020 = len(page2020.content.split())
    WC2021 = len(page2021.content.split())
    WC2022 = len(page2022.content.split())

    #print each years word count and add to list
    print("2010 word count: " + str(WC2010))
    lst.append(WC2010)
    print("2011 word count: " + str(WC2011))
    lst.append(WC2011)
    print("2012 word count: " + str(WC2012))
    lst.append(WC2012)
    print("2013 word count: " + str(WC2013))
    lst.append(WC2013)
    print("2014 word count: " + str(WC2014))
    lst.append(WC2014)
    print("2015 word count: " + str(WC2015))
    lst.append(WC2015)
    print("2016 word count: " + str(WC2016))
    lst.append(WC2016)
    print("2017 word count: " + str(WC2017))
    lst.append(WC2017)
    print("2018 word count: " + str(WC2018))
    lst.append(WC2018)
    print("2019 word count: " + str(WC2019))
    lst.append(WC2019)
    print("2020 word count: " + str(WC2020))
    lst.append(WC2020)
    print("2021 word count: " + str(WC2021))
    lst.append(WC2021)
    print("2022 word count: " + str(WC2022) + "\n")
    lst.append(WC2022)

    #add the word counts from the list to the value of the dictionary
    #this creates a dictionary with the year as the key and the word count as the value
    counter = 0
    for key in dic.keys():
        if(counter <= len(lst) -  1) and dic[key] != "Year" and dic[key] != "Word Count":
            dic[key] = lst[counter]
            counter += 1

    #prints the dictionary
    print(dic)
    #print("\n")

    #creates and excel file with our data
    makeSpreadsheet(dic)

def makeSpreadsheet(dic):
    #creates the xlxs
    fileName = "wordCounts.xlsx"
    fileData = [(key, value) for key, value in dic.items()]
    xlsxFile = openpyxl.Workbook()
    newWorkbook = xlsxFile["Sheet"]

    #add the data
    for row, (year, wordCount) in enumerate(fileData, 1):
        newWorkbook['A{}'.format(row)].value = year
        newWorkbook['B{}'.format(row)].value = wordCount

    #save
    xlsxFile.save("C:/Users/" + getpass.getuser() + "/Desktop/" + fileName)
    excelFileData = openpyxl.load_workbook("C:/Users/" + getpass.getuser() + "/Desktop/" + fileName)
    excelFileData.sheetnames
    sheetValues = excelFileData["Sheet"]

    #printAll(sheetValues)
    #creates the graph and adds it to the spreadsheet
    createGraph(sheetValues, newWorkbook, xlsxFile)

#use this if you want to print the values of the spreadsheet to the interpreter
def printAll(sheetValues):
    for row in sheetValues:
        for cell in row:
            print(cell.value)

def createGraph(sheetValues, newWorkbook, xlsxFile):
    #creats bar graph and adds title and axis data
    chart = BarChart()
    chart.type = "col"
    chart.title = "Years and Word Counts"
    chart.y_axis.title = "Word Counts"
    chart.x_axis.title = "Years"
    chart.style = 10

    #add data from spreadsheet to graph and save
    data = Reference(sheetValues, min_row = 2, max_row = 15, min_col = 1, max_col = 2)
    chart.add_data(data, from_rows = True, titles_from_data = True)
    newWorkbook.add_chart(chart, "A18")
    xlsxFile.save("C:/Users/" + getpass.getuser() + "/Desktop/" + "wordCounts.xlsx")